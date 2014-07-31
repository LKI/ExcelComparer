import sys
from openpyxl import *

def compareExcel(ename1, ename2):

    print '{0:10} : {1:40}'.format("Comparing", ename1) 
    print '{0:10} : {1:40}'.format(" ", ename2) 
    # A bool to verify if 2 xlsx is the same.
    fileSame = True

    # Load workbook, and get sheetname
    wb1 = load_workbook(filename = ename1)
    wb2 = load_workbook(filename = ename2)
    sn1 = wb1.get_sheet_names()
    sn2 = wb2.get_sheet_names()

    if (sn1 != sn2):
        print "Two file has different sheets."
        print ename1,"has sheet names:", sn1
        print ename2,"has sheet names:", sn2
    else:
        print "------------------------------------------------------------------------"
        print "|{0:17}|{2:10}|{1:40}|".format("Sheet Name", "Difference", "Location")
        sn = sn1
        for wsn in sn:
            # Get worksheet
            ws1 = wb1.get_sheet_by_name(name = wsn)
            ws2 = wb2.get_sheet_by_name(name = wsn)
            c = ws1.get_highest_column()
            r = ws1.get_highest_row()
            # This can be replace by ws1.get_dimension
            if ((ws2.get_highest_column()!= c) or (ws2.get_highest_row() != r)):
                print "|{0:17}|{2:10}|{1:40}|".format(wsn, "Dimension not the same!", "")
                fileSame = False
            else:
                #Compare every cell.
                flag = True
                for i in range(1,r+1):
                    for j in range(1,c+1):
                        c1 = ws1.cell(None,i,j)
                        c2 = ws2.cell(None,i,j)
                        if (c1):
                            if (c2):
                                if (c1.value != c2.value):
                                    if ((wsn == "Internal Info") and ((i == 4) and (j == 2)) or ((i == 5) and (j == 3))):
                                        continue
                                    print "|{0:17}|{2:10}|{1:40}|".format(wsn, c1.value, c1.coordinate)
                                    print "|{0:17}|{2:10}|{1:40}|".format(" ", c2.value, c2.coordinate)
                                    flag = False
                            else:
                                print "|{0:17}|{2:10}|{1:40}|".format(wsn, c1.value, c1.coordinate)
                                print "|{0:17}|{2:10}|{1:40}|".format(" ", "None")
                                flag = False
                        else:
                            if (c2):
                                print "|{0:17}|{2:10}|{1:40}|".format(wsn, "None")
                                print "|{0:17}|{2:10}|{1:40}|".format(" ", c2.value, c2.coordinate)
                                flag = False
                fileSame = fileSame and flag
        if fileSame:
            print "|{0:69}|".format("No differences.")
        print "------------------------------------------------------------------------"
        if fileSame:
            print '{0:10} : {1:40}'.format("Same file", ename1) 
            print '{0:10} : {1:40}'.format(" ", ename2) 

    print

if (len(sys.argv) != 3):
    print "2 filename args required.Must be .xlsx file."
else:
    f1 = sys.argv[1]
    f2 = sys.argv[2]
    compareExcel(f1,f2)


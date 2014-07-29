from openpyxl import *

def compareExcel(ename1, ename2):

    print "------------------------------------"
    print "Comparing",ename1,ename2
    fileSame = True

    wb1 = load_workbook(filename = ename1)
    wb2 = load_workbook(filename = ename2)
    sn1 = wb1.get_sheet_names()
    sn2 = wb2.get_sheet_names()

    if (sn1 != sn2):
        print "Two file has different sheets."
        print ename1,"has sheet names:", sn1
        print ename2,"has sheet names:", sn2
    else:
        sn = sn1
        for wsn in sn:
            ws1 = wb1.get_sheet_by_name(name = wsn)
            ws2 = wb2.get_sheet_by_name(name = wsn)
            c = ws1.get_highest_column()
            r = ws1.get_highest_row()
            if ((ws2.get_highest_column()!= c) or (ws2.get_highest_row() != r)):
                print "DIFFERDENT at SHEET-",wsn,": Rows or columns not the same!"
                fileSame = False
            else:
                flag = True
                for i in range(1,r+1):
                    for j in range(1,c+1):
                        c1 = ws1.cell(None,i,j)
                        c2 = ws2.cell(None,i,j)
                        if (c1):
                            if (c2):
                                if (c1.value != c2.value):
                                    if ((wsn == "Internal Info") and ((i == 4) and (j == 2)) or ((i == 5) and (j == 3))):
                                        continue
                                    print "DIFFERDENT_VALUE at SHEET-",wsn,": At (",i,",",j,")",
                                    print "diff FROM",c1.value,"TO",c2.value
                                    flag = False
                            else:
                                print "DIFFERDENT_TO_NONE at SHEET-",wsn,": At (",i,",",j,")"
                                print "diff FROM",c1.value
                                flag = False
                        else:
                            if (c2):
                                print "DIFFERDENT_TO_NONE at SHEET-",wsn,": At (",i,",",j,")"
                                print "diff FROM",c2.value
                                flag = False
                fileSame = fileSame and flag
        if fileSame:
            print "SAME_FILE:", ename1, ename2

    print "------------------------------------"
    print

compareExcel("r1\\ee\\H-Square_00061267.xlsx","r2\\ee\\H-Square_00061267.xlsx")
compareExcel("r1\\ee\\i2o_00061653.xlsx","r2\\ee\\i2o_00061653.xlsx")
compareExcel("r1\\ee\\Pexco_00061632.xlsx","r2\\ee\\Pexco_00061632.xlsx")
compareExcel("r1\\ee\\sai_00061574.xlsx","r2\\ee\\sai_00061574.xlsx")
compareExcel("r1\\ee\\teauru_00061560.xlsx","r2\\ee\\teauru_00061560.xlsx")
compareExcel("r1\\ee\\testcase8_1_00021702.xlsx","r2\\ee\\testcase8_1_00021702.xlsx")
compareExcel("r1\\ee\\testcase8_2_00021702.xlsx","r2\\ee\\testcase8_2_00021702.xlsx")
compareExcel("r1\\ee\\testcase8_3_00021702.xlsx","r2\\ee\\testcase8_3_00021702.xlsx")
compareExcel("r1\\ee\\vri_00061503.xlsx","r2\\ee\\vri_00061503.xlsx")
compareExcel("r1\\ee\\wpc_00061509.xlsx","r2\\ee\\wpc_00061509.xlsx")


compareExcel("r1\\se\\Alcoa Fastening Systems_00046757.xlsx","r2\\se\\Alcoa Fastening Systems_00046757.xlsx")
compareExcel("r1\\se\\AvonProductsCo.Ltd_00048758.xlsx","r2\\se\\AvonProductsCo.Ltd_00048758.xlsx")
compareExcel("r1\\se\\Bonar Floors SAS_00024080.xlsx","r2\\se\\Bonar Floors SAS_00024080.xlsx")
compareExcel("r1\\se\\Cerbona Rt_00048523.xlsx","r2\\se\\Cerbona Rt_00048523.xlsx")
compareExcel("r1\\se\\DanishCo_00099992.xlsx","r2\\se\\DanishCo_00099992.xlsx")
compareExcel("r1\\se\\Moog Berkeley Operations_00060869.xlsx","r2\\se\\Moog Berkeley Operations_00060869.xlsx")
compareExcel("r1\\se\\Perkins Engines Company Ltd_00002430.xlsx","r2\\se\\Perkins Engines Company Ltd_00002430.xlsx")
compareExcel("r1\\se\\PolskieLaboratoriumRadykalnychTe_00002544.xlsx","r2\\se\\PolskieLaboratoriumRadykalnychTe_00002544.xlsx")
compareExcel("r1\\se\\Rondo Food GmbH & Co. KG_00056043.xlsx","r2\\se\\Rondo Food GmbH & Co. KG_00056043.xlsx")
compareExcel("r1\\se\\Stabilit Servicios, SA de CV_00033252.xlsx","r2\\se\\Stabilit Servicios, SA de CV_00033252.xlsx")

